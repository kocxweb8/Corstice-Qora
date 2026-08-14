from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import io

def generate_excel(project, quantities, boq_items):
    wb = Workbook()
    # Sheet 1: Project Info
    ws1 = wb.active
    ws1.title = "Project Info"
    ws1.append(["Project", project.name])
    ws1.append(["Country", project.country])
    ws1.append(["Building Code", project.building_code])
    
    # Sheet 2: Quantities
    ws2 = wb.create_sheet("Quantities")
    ws2.append(["Item", "Gross", "Deduction", "Net", "Unit"])
    for key, val in quantities.items():
        ws2.append([key, val["gross"], val["deduction"], val["net"], val["unit"]])
    
    # Sheet 3: BOQ
    ws3 = wb.create_sheet("BOQ")
    ws3.append(["Description", "Qty", "Unit", "Rate", "Amount"])
    for item in boq_items:
        ws3.append([item.description, item.quantity, item.unit, item.rate, item.amount])
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf(project, quantities, boq_items):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.drawString(100, height-50, f"Corstice Estimate - {project.name}")
    c.drawString(100, height-70, f"Country: {project.country}")
    c.drawString(100, height-90, "Quantities Summary:")
    y = height-120
    for key, val in quantities.items():
        c.drawString(120, y, f"{key}: {val['net']} {val['unit']}")
        y -= 20
    c.drawString(100, y-20, "BOQ Items:")
    y -= 40
    for item in boq_items:
        c.drawString(120, y, f"{item.description}: {item.quantity} {item.unit} @ {item.rate} = {item.amount}")
        y -= 20
    c.save()
    buffer.seek(0)
    return buffer